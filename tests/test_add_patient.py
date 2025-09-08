import os
import time
import pytest
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright


# ============================
# CO
# ============================
ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_FILE_PATH = os.path.join(ROOT_DIR, "patient_details_updated.xlsx")
BASE_URL = "https://cx-dev-client.azurewebsites.net/login"


def load_excel_data():
    """Load patient details from Excel and return as list of dicts."""
    try:
        wb = load_workbook(EXCEL_FILE_PATH)
    except FileNotFoundError:
        pytest.skip(f"❌ Excel file not found at: {EXCEL_FILE_PATH}", allow_module_level=True)

    sheet = wb[wb.sheetnames[0]]
    headers = [cell.value.strip().replace(" ", "") for cell in sheet[1]]

    patients = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            break
        form_data = {headers[i]: row[i] for i in range(len(headers))}
        patients.append(form_data)
    return patients


def load_credentials():
    """Load login credentials from environment or fallback file."""
    username = "nick"
    password = "TEsplabs@123"

    if username and password:
        return username, password

    # Local fallback (only for dev machine)
    try:
        with open("credentials.txt", "r") as file:
            lines = file.read().splitlines()
            return lines[0].strip(), lines[1].strip()
    except FileNotFoundError:
        raise RuntimeError("❌ No credentials found in env or file")


@pytest.fixture(scope="session")
def browser_context():
    """Launch browser and login once per session."""
    username, password = load_credentials()

    with sync_playwright() as playwright:
        # Detect if running in CI
        is_ci = os.getenv("CI") == "true"

        browser = playwright.chromium.launch(
            headless=is_ci,
            slow_mo=1000 if not is_ci else 0,
            args=["--start-maximized"] if not is_ci else []
        )

        # Context settings differ for CI vs Local
        if is_ci:
            context = browser.new_context(
                viewport={"width": 1366, "height": 768}  # safe CI resolution
            )
        else:
            context = browser.new_context(no_viewport=True)  # full window locally

        page = context.new_page()

        # Open login page and wait until fully loaded
        page.goto(BASE_URL, wait_until="networkidle")
        page.wait_for_selector("#login_username", timeout=15000)

        # Fill credentials
        page.fill("#login_username", username)
        page.fill("#login_password", password)
         page.fill("#login_password", password)
        # Wait for login button enabled and click
        page.wait_for_selector("#btn_login:enabled", timeout=15000)
        page.click("#btn_login")
        time.sleep(30)
        # Allow time for navigation
        page.wait_for_load_state("networkidle")
        page.wait_for_selector("#homeaddpatient", timeout=30000)

        yield page
        browser.close()


@pytest.mark.parametrize("form_data", load_excel_data())
def test_add_patient(browser_context, form_data):
    """Add patient using details from Excel."""
    page = browser_context

    if not form_data.get("Firstname") or not form_data.get("Lastname"):
        pytest.skip("⚠️ Skipping row due to missing Firstname/Lastname")

    # Navigate to add patient
    page.click("#homeaddpatient")
    time.sleep(45)
    page.wait_for_load_state("networkidle")
    page.wait_for_selector("#addPatientFirstname", timeout=10000)

    # Fill patient form
    page.fill("#addPatientFirstname", str(form_data.get("Firstname", "")))
    page.fill("#addPatientMiddlename", str(form_data.get("Middlename", "")))
    page.fill("#addPatientlastname", str(form_data.get("Lastname", "")))
    page.fill("#addPatientemail", str(form_data.get("Email", "")))
    page.fill("#addPatientMobile", str(form_data.get("MobileNumber", "")))
    page.fill("#addPatientDOB", "1997-10-01")

    page.select_option("#addPatientheight", value=str(form_data.get("Height", "")))
    page.select_option("#addPatientGender", value="M")
    page.select_option("#addPatientClinicName", value="8")

    page.fill("#addPatientAddressLine1", str(form_data.get("AddressLine1", "")))
    page.select_option("#addPatientState", value="1399")
    page.select_option("#addPatientCity", label="Bellefonte")
    page.wait_for_selector("#addPatientZipcode", state="visible")
    page.fill("#addPatientZipcode", "100")
    page.select_option("#addPatientTimezone", value="1")
    page.fill("#addPatientEmergencyContact1", value="Andria")
    page.select_option("#addPatientRelation1", index=1)
    page.fill("xpath=//*[@id='addPatientRelation1_mobile']", value="1234567890")
    page.fill("#addPatientAdditionalNotes", str(form_data.get("Notes", "")))

    # Submit draft
    page.click("#btnaddPatientDraftSubmit")
    locator = page.locator("h4.subheading-dialog")
    locator.wait_for(state="visible", timeout=5000)
    assert locator.inner_text() == "New Patient added Successfully!"

    # Program & Vital
    program_name = str(form_data.get("ProgramName", "")).strip()
    vital_condition = str(form_data.get("VitalCondition", "")).strip()

    page.click("#addPatient")

    if program_name:
        page.select_option("#addPatientProgramName", label=program_name)
        page.fill("#addPatientStartDate", "2025-06-20")

    if vital_condition:
        page.wait_for_selector("#addPatientvitalChange .mat-mdc-select-trigger", timeout=30000)
        page.click("#addPatientvitalChange .mat-mdc-select-trigger")
        page.wait_for_selector("mat-option span.mdc-list-item__primary-text", timeout=30000)
        page.get_by_text(vital_condition, exact=True).click()

    # Diagnosis
    diagnosis = str(form_data.get("Diagnosis", "")).strip()
    if diagnosis:
        page.get_by_role("button", name="Program Info").click(force=True)
        page.click("input[placeholder='Select Diagnosis']")
        page.get_by_text(diagnosis, exact=True).click()

    # Confirm
    page.click("#btnaddPatientConfirmSubmit")
    page.locator("button.btn_save", has_text="CONFIRM").click()

    # Verify prescribed
    page.get_by_role("button", name="VIEW PATIENT").click()
    text = page.locator("span.status_display.patient_prescribed").first.text_content()
    assert text.strip() == "Prescribed"

    # Go back home for next iteration
    page.locator("div.menu-items:has(h3.menu-title:has-text('Home'))").click(force=True)
    page.wait_for_load_state("networkidle")
